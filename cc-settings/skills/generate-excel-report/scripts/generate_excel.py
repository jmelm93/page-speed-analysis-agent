#!/usr/bin/env python3
"""Generate Excel report from page speed analysis data and save locally.

Creates a 4-sheet Excel workbook with:
1. Summary - All URL scores at a glance
2. Core Web Vitals - Full metrics breakdown (lab + field)
3. Network Analysis - Resource details by type
4. Opportunities - All PSI recommendations with savings

Usage:
    python generate_excel.py --data ./collected_data.json
    python generate_excel.py --data ./collected_data.json --output-dir ./output
"""
import argparse
import io
import json
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import ColorScaleRule


# --- Inlined utility: sanitize_for_excel ---

def sanitize_for_excel(value: Any, max_length: int = 32767) -> str:
    """Sanitize content for safe Excel cell writing.

    Removes illegal XML control characters that cause
    openpyxl.utils.exceptions.IllegalCharacterError.
    """
    if value is None:
        return ""
    try:
        content = str(value)
        content = re.sub(r'[\x00-\x08\x0B-\x0C\x0E-\x1F]', '', content)
        content = content.replace('\uFFFE', '').replace('\uFFFF', '')
        if len(content) > max_length:
            content = content[:max_length - 20]
            last_space = content.rfind(' ')
            if last_space > max_length * 0.8:
                content = content[:last_space] + "... [truncated]"
            else:
                content = content + "... [truncated]"
        return content
    except Exception:
        return "[Content sanitization failed]"


# --- Styling constants ---

HEADER_FONT = Font(bold=True)
HEADER_FILL = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
CENTER_ALIGN = Alignment(horizontal="center")


def create_summary_sheet(wb: openpyxl.Workbook, data: dict) -> None:
    """Create Summary sheet with overview of all URLs and scores."""
    ws = wb.create_sheet("Summary", 0)

    headers = [
        "URL",
        "Mobile Score",
        "Desktop Score",
        "LCP Status",
        "INP Status",
        "CLS Status",
        "Top Issue",
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN

    psi_results = data.get("psi_results", [])

    for row_idx, psi in enumerate(psi_results, 2):
        url = psi.get("url", "")
        strategies = psi.get("strategies", {})

        mobile = strategies.get("mobile", {})
        desktop = strategies.get("desktop", {})

        mobile_cwv = mobile.get("core_web_vitals", {})

        mobile_opps = mobile.get("opportunities", [])
        top_issue = mobile_opps[0].get("title", "") if mobile_opps else ""

        ws.cell(row=row_idx, column=1, value=sanitize_for_excel(url))
        ws.cell(row=row_idx, column=2, value=mobile.get("performance_score", ""))
        ws.cell(row=row_idx, column=3, value=desktop.get("performance_score", ""))
        ws.cell(row=row_idx, column=4, value=_get_status_text(mobile_cwv.get("lcp", {})))
        ws.cell(row=row_idx, column=5, value=_get_status_text(mobile_cwv.get("inp", {})))
        ws.cell(row=row_idx, column=6, value=_get_status_text(mobile_cwv.get("cls", {})))
        ws.cell(row=row_idx, column=7, value=sanitize_for_excel(top_issue))

    if psi_results:
        row_count = len(psi_results) + 1
        color_scale = ColorScaleRule(
            start_type="min", start_color="F8696B",
            mid_type="percentile", mid_value=50, mid_color="FFEB84",
            end_type="max", end_color="63BE7B"
        )
        ws.conditional_formatting.add(f"B2:C{row_count}", color_scale)

    ws.freeze_panes = "A2"

    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 50


def create_core_web_vitals_sheet(wb: openpyxl.Workbook, data: dict) -> None:
    """Create Core Web Vitals sheet with full metrics breakdown."""
    ws = wb.create_sheet("Core Web Vitals", 1)

    headers = [
        "URL", "Strategy", "Score",
        "LCP Lab", "LCP Field", "INP Lab", "INP Field",
        "CLS Lab", "CLS Field", "FCP", "TTFB", "TBT",
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN

    psi_results = data.get("psi_results", [])
    row_idx = 2

    for psi in psi_results:
        url = psi.get("url", "")
        strategies = psi.get("strategies", {})

        for strategy_name, strategy_data in strategies.items():
            cwv = strategy_data.get("core_web_vitals", {})
            field = strategy_data.get("field_metrics", {}) or {}

            lcp = cwv.get("lcp", {})
            inp = cwv.get("inp", {})
            cls = cwv.get("cls", {})
            fcp = cwv.get("fcp", {})
            ttfb = cwv.get("ttfb", {})
            tbt = cwv.get("tbt", {})

            ws.cell(row=row_idx, column=1, value=sanitize_for_excel(url))
            ws.cell(row=row_idx, column=2, value=strategy_name.capitalize())
            ws.cell(row=row_idx, column=3, value=strategy_data.get("performance_score", ""))
            ws.cell(row=row_idx, column=4, value=_format_metric(lcp))
            ws.cell(row=row_idx, column=5, value=_format_field_metric(field, "lcp_p75", "s"))
            ws.cell(row=row_idx, column=6, value=_format_metric(inp))
            ws.cell(row=row_idx, column=7, value=_format_field_metric(field, "inp_p75", "ms"))
            ws.cell(row=row_idx, column=8, value=_format_metric(cls))
            ws.cell(row=row_idx, column=9, value=_format_field_metric(field, "cls_p75", ""))
            ws.cell(row=row_idx, column=10, value=_format_metric(fcp))
            ws.cell(row=row_idx, column=11, value=_format_metric(ttfb))
            ws.cell(row=row_idx, column=12, value=_format_metric(tbt))

            row_idx += 1

    ws.freeze_panes = "A2"

    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 8
    for col in "DEFGHIJKL":
        ws.column_dimensions[col].width = 12


def create_network_analysis_sheet(wb: openpyxl.Workbook, data: dict) -> None:
    """Create Network Analysis sheet with resource breakdown."""
    ws = wb.create_sheet("Network Analysis", 2)

    headers = [
        "URL", "Total Requests", "Transfer (MB)",
        "Scripts (count)", "Scripts (KB)",
        "Images (count)", "Images (KB)",
        "CSS (count)", "CSS (KB)",
        "Fonts (count)", "Fonts (KB)",
        "Largest Resource",
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN

    network_results = data.get("network_results", [])

    for row_idx, network in enumerate(network_results, 2):
        url = network.get("url", "")
        # Handle both flat and nested (summary) formats from capture_network.py
        summary = network.get("summary", {})
        by_type = network.get("by_type") or summary.get("by_type", {})
        total_requests = network.get("total_requests") or summary.get("total_requests", 0)
        total_bytes = network.get("total_transfer_bytes") or summary.get("total_transfer_bytes", 0)
        largest = network.get("largest_resources", [])

        scripts = by_type.get("script", {})
        images = by_type.get("image", {})
        css = by_type.get("stylesheet", {})
        fonts = by_type.get("font", {})

        largest_name = ""
        if largest:
            resource = largest[0]
            size_kb = resource.get("size_bytes", 0) / 1024
            largest_name = f"{resource.get('url', '').split('/')[-1][:30]} ({size_kb:.0f}KB)"

        ws.cell(row=row_idx, column=1, value=sanitize_for_excel(url))
        ws.cell(row=row_idx, column=2, value=total_requests)
        ws.cell(row=row_idx, column=3, value=round(total_bytes / 1024 / 1024, 2))
        ws.cell(row=row_idx, column=4, value=scripts.get("count", 0))
        ws.cell(row=row_idx, column=5, value=round(scripts.get("bytes", 0) / 1024, 0))
        ws.cell(row=row_idx, column=6, value=images.get("count", 0))
        ws.cell(row=row_idx, column=7, value=round(images.get("bytes", 0) / 1024, 0))
        ws.cell(row=row_idx, column=8, value=css.get("count", 0))
        ws.cell(row=row_idx, column=9, value=round(css.get("bytes", 0) / 1024, 0))
        ws.cell(row=row_idx, column=10, value=fonts.get("count", 0))
        ws.cell(row=row_idx, column=11, value=round(fonts.get("bytes", 0) / 1024, 0))
        ws.cell(row=row_idx, column=12, value=sanitize_for_excel(largest_name))

    ws.freeze_panes = "A2"

    ws.column_dimensions["A"].width = 60
    for col in "BCDEFGHIJKL":
        ws.column_dimensions[col].width = 14


def create_opportunities_sheet(wb: openpyxl.Workbook, data: dict) -> None:
    """Create Opportunities sheet with all PSI recommendations."""
    ws = wb.create_sheet("Opportunities", 3)

    headers = [
        "URL", "Strategy", "Opportunity",
        "Est. Savings (ms)", "Est. Savings (KB)", "Description",
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN

    psi_results = data.get("psi_results", [])
    row_idx = 2

    for psi in psi_results:
        url = psi.get("url", "")
        strategies = psi.get("strategies", {})

        for strategy_name, strategy_data in strategies.items():
            opportunities = strategy_data.get("opportunities", [])

            for opp in opportunities:
                ws.cell(row=row_idx, column=1, value=sanitize_for_excel(url))
                ws.cell(row=row_idx, column=2, value=strategy_name.capitalize())
                ws.cell(row=row_idx, column=3, value=sanitize_for_excel(opp.get("title", "")))
                ws.cell(row=row_idx, column=4, value=opp.get("savings_ms", ""))

                savings_bytes = opp.get("savings_bytes", 0)
                savings_kb = round(savings_bytes / 1024, 1) if savings_bytes else ""
                ws.cell(row=row_idx, column=5, value=savings_kb)

                desc = opp.get("description", "")
                if len(desc) > 200:
                    desc = desc[:197] + "..."
                ws.cell(row=row_idx, column=6, value=sanitize_for_excel(desc))

                row_idx += 1

    ws.freeze_panes = "A2"

    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 80


def _get_status_text(metric: dict) -> str:
    """Get human-readable status from metric dict."""
    status = metric.get("status", "")
    if status == "good":
        return "Good"
    elif status == "needs_improvement":
        return "Needs Improvement"
    elif status == "poor":
        return "Poor"
    return ""


def _format_metric(metric: dict) -> str:
    """Format a metric value with unit."""
    if not metric:
        return ""
    value = metric.get("value", "")
    unit = metric.get("unit", "")
    if value == "":
        return ""
    return f"{value}{unit}"


def _format_field_metric(field: dict, key: str, unit: str) -> str:
    """Format a field metric value."""
    if not field or key not in field:
        return "N/A"
    value = field[key]
    if value is None:
        return "N/A"
    return f"{value}{unit}"


def generate_excel(data: dict, output_dir: str = "./output", job_id: str | None = None) -> dict:
    """Generate Excel workbook and save locally.

    Args:
        data: Collected page speed data (psi_results, crux_results, network_results)
        output_dir: Directory to save the Excel file
        job_id: Optional job ID for filename prefix

    Returns:
        Dict with success status, file_path, filename, sheets, url_count, generated_at
    """
    # Create workbook
    wb = openpyxl.Workbook()

    # Create all 4 sheets
    create_summary_sheet(wb, data)
    create_core_web_vitals_sheet(wb, data)
    create_network_analysis_sheet(wb, data)
    create_opportunities_sheet(wb, data)

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Build filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    prefix = f"{job_id}_" if job_id else ""
    filename = f"page_speed_analysis_{prefix}{timestamp}.xlsx"

    # Ensure output directory exists
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    # Save locally
    file_path = output_path / filename
    wb.save(str(file_path))

    # Count URLs
    url_count = len(data.get("urls", [])) or len(data.get("psi_results", []))

    return {
        "success": True,
        "file_path": str(file_path),
        "filename": filename,
        "sheets": ["Summary", "Core Web Vitals", "Network Analysis", "Opportunities"],
        "url_count": url_count,
        "generated_at": datetime.now().isoformat() + "Z",
    }


def main():
    """Main entry point for the Excel generation skill."""
    parser = argparse.ArgumentParser(
        description="Generate Excel report from page speed analysis data"
    )
    parser.add_argument(
        "--data",
        required=True,
        help="Path to JSON file with collected PSI/CrUX/network data"
    )
    parser.add_argument(
        "--output-dir",
        default="./output",
        help="Directory to save the Excel file (default: ./output)"
    )
    parser.add_argument(
        "--job-id",
        help="Optional job ID for filename prefix"
    )
    args = parser.parse_args()

    # Read data file
    try:
        with open(args.data, "r") as f:
            data = json.load(f)
    except FileNotFoundError:
        result = {
            "success": False,
            "error": f"Failed to read data file: File not found: {args.data}"
        }
        print(json.dumps(result, indent=2))
        return 1
    except json.JSONDecodeError as e:
        result = {
            "success": False,
            "error": f"Failed to parse JSON data: {str(e)}"
        }
        print(json.dumps(result, indent=2))
        return 1

    # Validate data format
    psi_results = data.get("psi_results", [])
    network_results = data.get("network_results", [])

    if not psi_results:
        # Check if file paths were provided instead of inline data
        if data.get("psi_files"):
            result = {
                "success": False,
                "error": "Invalid data format: 'psi_files' found but 'psi_results' is missing. "
                         "The collected_data.json must contain actual PSI data inline in 'psi_results', "
                         "not file path references. Read each PSI JSON file and include the data directly."
            }
            print(json.dumps(result, indent=2))
            return 1
        result = {
            "success": False,
            "error": "No PSI results found in data. The 'psi_results' array is empty. "
                     "Ensure collected_data.json contains PSI data before running Excel generation."
        }
        print(json.dumps(result, indent=2))
        return 1

    # Generate Excel
    try:
        result = generate_excel(data, output_dir=args.output_dir, job_id=args.job_id)
        print(json.dumps(result, indent=2))
        return 0
    except Exception as e:
        result = {
            "success": False,
            "error": f"Failed to generate Excel report: {str(e)}"
        }
        print(json.dumps(result, indent=2))
        return 1


if __name__ == "__main__":
    sys.exit(main())
